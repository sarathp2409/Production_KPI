"""
Generate Excel Templates with Dropdown Validation for Historical Pagination Data Collection

This script generates per-plant .xlsx files with:
- Dropdown validation for all CHOICE and Boolean fields
- Plant-specific GNP columns (only relevant publications per plant)
- Color-coded headers (required vs optional)
- Pre-filled plant_id and plant_name
- Reference sheet with all valid values

Usage:
    python pagination_dash/generate_xlsx_templates.py

Requirements:
    pip install openpyxl
"""

import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter


import json

# Plant configurations
PLANT_DETAILS = [
    {"id": "ODBCAIR30", "name": "Airoli", "display_name": "Airoli"},
    {"id": "ACTIVE42", "name": "Baroda", "display_name": "Baroda"},
    {"id": "ODBCAIR32", "name": "Pune", "display_name": "Bhosari"},
    {"id": "MAIN34", "name": "Bangalore", "display_name": "Bommasandra"},
    {"id": "MAIN38", "name": "Nagpur", "display_name": "Butibori"},
    {"id": "MAIN35", "name": "Chennai", "display_name": "Chemmencherry"},
    {"id": "MAIN40", "name": "Lucknow", "display_name": "Chinhat"},
    {"id": "ODBCAIR28", "name": "Kandivali", "display_name": "Kandivali"},
    {"id": "ACTIVE44", "name": "Manesar", "display_name": "Manesar"},
    {"id": "MAIN36", "name": "Hyderabad", "display_name": "Nacharam"},
    {"id": "MAIN41", "name": "Sahibabad", "display_name": "Sahibabad"},
    {"id": "ODBCAIR33", "name": "Kolkata", "display_name": "Saltlake"},
    {"id": "ACTIVE43", "name": "Trivandrum", "display_name": "Trivandrum"},
    {"id": "MAIN37", "name": "Ahmedabad", "display_name": "Vejalpur"},
]

# Load competitor information
def load_competitor_info():
    """Load competitor information from JSON file"""
    try:
        comp_file = os.path.join(os.path.dirname(__file__), 'comp_info.json')
        with open(comp_file, 'r') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Could not load competitor info: {e}")
        return {}

COMPETITOR_INFO = load_competitor_info()

# GNP publications per plant
PLANT_GNP_PUBLICATIONS = {
    "Airoli": ["Mirror", "MT", "TIMS"],
    "Baroda": [],
    "Bhosari": ["ET", "MT", "TIMS"],
    "Bommasandra": ["ET", "TIMS", "Mirror"],
    "Butibori": [],
    "Chemmencherry": ["ET", "TIMS"],
    "Chinhat": ["ET", "NBT", "TIMS"],
    "Kandivali": ["ET", "MT", "NBT", "Mirror", "TIMS"],
    "Manesar": ["ET", "NBT", "TIMS"],
    "Nacharam": ["ET", "TIMS"],
    "Sahibabad": ["ET", "NBT", "TIMS"],
    "Saltlake": ["ET", "TIMS"],
    "Trivandrum": [],
    "Vejalpur": ["ET", "TIMS"],
}

# Valid choices for dropdowns
BOOK_NAME_CHOICES = [
    'TOI B-1',
    'TOI B-2',
    'TIMS B-1',
    'TIMS B-2',
    'TOI Supp-1',
    'TOI Supp-2',
    'TOI Supp-3',
]

GNP_TYPE_CHOICES = [
    '',  # Empty option first
    '2 Pg Jkt',
    '4 Pg Jkt - 2 diff clients',
    '4 Pg Vantage',
    '4 Pg Power Jkt',
    '8 Pg Super-pano',
    '6 Pg GNP',
    '8 Pg GNP',
]

INNOVATION_CHOICES = [
    '',  # Empty option first
    'Center Panorama',
    'Bookmark',
    'Seamless Panorama',
    'Front Vantage',
    'Reverse Flap',
    'Fragrance',
    'French-Window',
    'Power Jacket',
    'Others',
]

YES_NO_CHOICES = ['', 'Yes', 'No']

# Styles
HEADER_FILL_REQUIRED = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FILL_OPTIONAL = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
HEADER_FILL_DROPDOWN = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
HEADER_FILL_LOCKED = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Number of data rows to prepare (for dropdown validation range)
DATA_ROWS = 1000


def create_dropdown_validation(choices, allow_blank=True, error_title="Invalid Value"):
    """Create a DataValidation object for dropdown"""
    # For long lists, we need to handle differently
    formula = '"' + ','.join(choices) + '"'

    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
        showDropDown=False,  # False = show dropdown arrow (counterintuitive!)
        showErrorMessage=True,
        errorTitle=error_title,
        error="Please select a value from the dropdown list."
    )
    return dv


def get_inhouse_columns(plant_display_name):
    """Get column definitions for in-house data sheet based on plant"""
    gnp_pubs = PLANT_GNP_PUBLICATIONS.get(plant_display_name, [])
    has_et = 'ET' in gnp_pubs
    has_mt = 'MT' in gnp_pubs
    has_mirror = 'Mirror' in gnp_pubs
    has_nbt = 'NBT' in gnp_pubs
    has_tims = 'TIMS' in gnp_pubs

    # Base columns (always present)
    columns = [
        {'name': 'plant_id', 'type': 'locked', 'required': True, 'width': 12},
        {'name': 'plant_name', 'type': 'locked', 'required': True, 'width': 15},
        {'name': 'issue_date', 'type': 'date', 'required': True, 'width': 12},
        {'name': 'book_name', 'type': 'dropdown', 'choices': BOOK_NAME_CHOICES, 'required': True, 'width': 12},
        {'name': 'snp_pages', 'type': 'number', 'required': True, 'width': 10},
        {'name': 'gnp_pages_type', 'type': 'dropdown', 'choices': GNP_TYPE_CHOICES, 'required': False, 'width': 25},
        {'name': 'number_of_gnp_jackets', 'type': 'number', 'required': False, 'width': 18},
        {'name': 'TOI Book-2 balloon_printed', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 25},
        {'name': 'TOI Book-2 has_masthead', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 25},
        {'name': 'extra_miles_toi', 'type': 'text', 'required': False, 'width': 25},
        {'name': 'extra_miles_others', 'type': 'text', 'required': False, 'width': 25},
        {'name': 'first_time_execution_info', 'type': 'text', 'required': False, 'width': 25},
    ]

    # Add et_had_2_books right after first_time_execution_info (if plant has ET)
    if has_et:
        columns.append({'name': 'et_had_2_books', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 18})

    # Add GNP columns based on plant configuration
    if has_et:
        columns.append({'name': 'gnp_et', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 10})
        columns.append({'name': 'gnp_et_b1', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 12})
        columns.append({'name': 'gnp_et_b2', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 12})

    if has_mt:
        columns.append({'name': 'gnp_mt', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 10})

    if has_mirror:
        columns.append({'name': 'gnp_mirror', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 12})

    if has_nbt:
        columns.append({'name': 'gnp_nbt', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 10})

    if has_tims:
        columns.append({'name': 'gnp_tims', 'type': 'dropdown', 'choices': YES_NO_CHOICES, 'required': False, 'width': 12})

    # Add GNP remarks at the end if plant has any GNP
    if gnp_pubs:
        columns.append({'name': 'gnp_remarks', 'type': 'text', 'required': False, 'width': 25})

    return columns


def get_competitor_columns(plant_display_name):
    """Get column definitions for competitor data sheet"""
    # Get plant-specific competitors
    plant_key = plant_display_name.lower()
    competitor_choices = COMPETITOR_INFO.get(plant_key, [])

    # Competitor name field - use dropdown if competitors are defined, otherwise text
    if competitor_choices:
        competitor_field = {'name': 'competitor_name', 'type': 'dropdown', 'choices': competitor_choices, 'required': True, 'width': 25}
    else:
        competitor_field = {'name': 'competitor_name', 'type': 'text', 'required': True, 'width': 25}

    return [
        {'name': 'plant_id', 'type': 'locked', 'required': True, 'width': 12},
        {'name': 'plant_name', 'type': 'locked', 'required': True, 'width': 15},
        {'name': 'issue_date', 'type': 'date', 'required': True, 'width': 12},
        competitor_field,
        {'name': 'main_snp_pages', 'type': 'number', 'required': False, 'width': 15},
        {'name': 'main_gnp_pages_type', 'type': 'dropdown', 'choices': GNP_TYPE_CHOICES, 'required': False, 'width': 25},
        {'name': 'main_number_of_gnp_jacket', 'type': 'number', 'required': False, 'width': 22},
        {'name': 'main_number_of_books', 'type': 'number', 'required': False, 'width': 18},
        {'name': 'supp_snp_pages', 'type': 'number', 'required': False, 'width': 15},
        {'name': 'supp_gnp_pages_type', 'type': 'dropdown', 'choices': GNP_TYPE_CHOICES, 'required': False, 'width': 25},
        {'name': 'supp_number_of_gnp_jacket', 'type': 'number', 'required': False, 'width': 22},
        {'name': 'supp_number_of_books', 'type': 'number', 'required': False, 'width': 18},
        {'name': 'innovation_1', 'type': 'dropdown', 'choices': INNOVATION_CHOICES, 'required': False, 'width': 20},
        {'name': 'innovation_2', 'type': 'dropdown', 'choices': INNOVATION_CHOICES, 'required': False, 'width': 20},
        {'name': 'innovation_3', 'type': 'dropdown', 'choices': INNOVATION_CHOICES, 'required': False, 'width': 20},
        {'name': 'comment', 'type': 'text', 'required': False, 'width': 30},
    ]


def create_inhouse_sheet(wb, plant_id, plant_name, plant_display_name):
    """Create the In-House Data sheet with dropdowns"""
    ws = wb.create_sheet("In-House Data", 0)

    columns = get_inhouse_columns(plant_display_name)

    # Create headers
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def['name'])
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

        # Color based on column type
        if col_def['type'] == 'locked':
            cell.fill = HEADER_FILL_LOCKED
        elif col_def['type'] == 'dropdown':
            cell.fill = HEADER_FILL_DROPDOWN
        elif col_def['required']:
            cell.fill = HEADER_FILL_REQUIRED
        else:
            cell.fill = HEADER_FILL_OPTIONAL

        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def['width']

    # Pre-fill plant_id and plant_name for data rows
    for row in range(2, DATA_ROWS + 2):
        ws.cell(row=row, column=1, value=plant_id)
        ws.cell(row=row, column=2, value=plant_display_name)

    # Add dropdown validations
    for col_idx, col_def in enumerate(columns, start=1):
        if col_def['type'] == 'dropdown':
            col_letter = get_column_letter(col_idx)
            dv = create_dropdown_validation(
                col_def['choices'],
                allow_blank=not col_def['required'],
                error_title=f"Invalid {col_def['name']}"
            )
            # Apply to column range
            dv.add(f'{col_letter}2:{col_letter}{DATA_ROWS + 1}')
            ws.add_data_validation(dv)

    # Add example rows
    example_date = "2023-01-01"
    example_rows = [
        {'book_name': 'TOI B-1', 'snp_pages': 24, 'gnp_pages_type': '4 Pg Vantage', 'number_of_gnp_jackets': 1},
        {'book_name': 'TOI B-2', 'snp_pages': 16, 'TOI Book-2 balloon_printed': 'Yes', 'TOI Book-2 has_masthead': 'Yes'},
        {'book_name': 'TIMS B-1', 'snp_pages': 20},
    ]

    for row_offset, example in enumerate(example_rows):
        row_num = 2 + row_offset
        for col_idx, col_def in enumerate(columns, start=1):
            col_name = col_def['name']
            if col_name == 'issue_date':
                ws.cell(row=row_num, column=col_idx, value=example_date)
            elif col_name in example:
                ws.cell(row=row_num, column=col_idx, value=example[col_name])

    # Freeze header row
    ws.freeze_panes = 'A2'

    return ws


def create_competitor_sheet(wb, plant_id, plant_name, plant_display_name):
    """Create the Competitor Data sheet with dropdowns"""
    ws = wb.create_sheet("Competitor Data", 1)

    columns = get_competitor_columns(plant_display_name)

    # Create headers
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def['name'])
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

        # Color based on column type
        if col_def['type'] == 'locked':
            cell.fill = HEADER_FILL_LOCKED
        elif col_def['type'] == 'dropdown':
            cell.fill = HEADER_FILL_DROPDOWN
        elif col_def['required']:
            cell.fill = HEADER_FILL_REQUIRED
        else:
            cell.fill = HEADER_FILL_OPTIONAL

        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def['width']

    # Pre-fill plant_id and plant_name for data rows
    for row in range(2, DATA_ROWS + 2):
        ws.cell(row=row, column=1, value=plant_id)
        ws.cell(row=row, column=2, value=plant_display_name)

    # Add dropdown validations
    for col_idx, col_def in enumerate(columns, start=1):
        if col_def['type'] == 'dropdown':
            col_letter = get_column_letter(col_idx)
            dv = create_dropdown_validation(
                col_def['choices'],
                allow_blank=not col_def['required'],
                error_title=f"Invalid {col_def['name']}"
            )
            # Apply to column range
            dv.add(f'{col_letter}2:{col_letter}{DATA_ROWS + 1}')
            ws.add_data_validation(dv)

    # Add example row
    # Use plant-specific competitor if available
    plant_key = plant_display_name.lower()
    competitor_choices = COMPETITOR_INFO.get(plant_key, [])
    sample_competitor = competitor_choices[0] if competitor_choices else 'Sample Competitor'

    example_row = {
        'issue_date': '2023-01-01',
        'competitor_name': sample_competitor,
        'main_snp_pages': 24,
        'main_gnp_pages_type': '4 Pg Vantage',
        'main_number_of_gnp_jacket': 1,
        'main_number_of_books': 2,
        'supp_snp_pages': 8,
        'supp_number_of_books': 1,
        'innovation_1': 'Center Panorama',
    }

    row_num = 2
    for col_idx, col_def in enumerate(columns, start=1):
        col_name = col_def['name']
        if col_name in example_row:
            ws.cell(row=row_num, column=col_idx, value=example_row[col_name])

    # Freeze header row
    ws.freeze_panes = 'A2'

    return ws


def create_reference_sheet(wb, plant_display_name):
    """Create the Reference sheet with all valid values"""
    ws = wb.create_sheet("Reference", 2)

    gnp_pubs = PLANT_GNP_PUBLICATIONS.get(plant_display_name, [])

    # Title
    ws['A1'] = "REFERENCE: Valid Values for Dropdown Fields"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # Plant info
    ws['A3'] = f"Plant: {plant_display_name}"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"GNP Publications: {', '.join(gnp_pubs) if gnp_pubs else 'None'}"

    # Book Names
    ws['A6'] = "Book Names (book_name)"
    ws['A6'].font = Font(bold=True)
    ws['A6'].fill = HEADER_FILL_DROPDOWN
    ws['A6'].font = HEADER_FONT
    for i, choice in enumerate(BOOK_NAME_CHOICES, start=7):
        ws[f'A{i}'] = choice

    # GNP Types
    ws['B6'] = "GNP Page Types"
    ws['B6'].font = Font(bold=True)
    ws['B6'].fill = HEADER_FILL_DROPDOWN
    ws['B6'].font = HEADER_FONT
    for i, choice in enumerate(GNP_TYPE_CHOICES[1:], start=7):  # Skip empty
        ws[f'B{i}'] = choice

    # Innovation Types
    ws['C6'] = "Innovation Types"
    ws['C6'].font = Font(bold=True)
    ws['C6'].fill = HEADER_FILL_DROPDOWN
    ws['C6'].font = HEADER_FONT
    for i, choice in enumerate(INNOVATION_CHOICES[1:], start=7):  # Skip empty
        ws[f'C{i}'] = choice

    # Yes/No
    ws['D6'] = "Yes/No Fields"
    ws['D6'].font = Font(bold=True)
    ws['D6'].fill = HEADER_FILL_DROPDOWN
    ws['D6'].font = HEADER_FONT
    ws['D7'] = "Yes"
    ws['D8'] = "No"
    ws['D9'] = "(empty = No)"

    # Instructions
    ws['A20'] = "INSTRUCTIONS"
    ws['A20'].font = Font(bold=True, size=12)

    instructions = [
        "1. Fill data in 'In-House Data' sheet first (one row per book per date)",
        "2. Fill 'Competitor Data' sheet after in-house data",
        "3. Use dropdown arrows to select values - DO NOT type manually",
        "4. Date format: YYYY-MM-DD (e.g., 2023-01-15)",
        "5. balloon_printed and has_masthead only apply to TOI B-2 and TIMS B-2",
        "6. GNP flags (gnp_et, gnp_mt, etc.) only need to be filled in first row of each date",
        "7. Delete example rows before submitting",
        "",
        "HEADER COLORS:",
        "  Gray = Pre-filled (do not edit)",
        "  Dark Blue = Required field",
        "  Light Blue = Optional field",
        "  Green = Dropdown field (use dropdown to select)",
    ]

    for i, instruction in enumerate(instructions, start=21):
        ws[f'A{i}'] = instruction

    # Set column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15

    return ws


def generate_plant_template(plant, output_dir):
    """Generate Excel template for a single plant"""
    plant_id = plant['id']
    plant_name = plant['name']
    plant_display_name = plant['display_name']

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create sheets
    create_inhouse_sheet(wb, plant_id, plant_name, plant_display_name)
    create_competitor_sheet(wb, plant_id, plant_name, plant_display_name)
    create_reference_sheet(wb, plant_display_name)

    # Save workbook
    filename = f"{plant_display_name}_pagination_data.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath


def generate_all_templates(output_dir):
    """Generate Excel templates for all plants"""
    os.makedirs(output_dir, exist_ok=True)

    print("\n" + "=" * 80)
    print("GENERATING EXCEL TEMPLATES WITH DROPDOWN VALIDATION")
    print("=" * 80 + "\n")

    generated_files = []
    for plant in PLANT_DETAILS:
        filepath = generate_plant_template(plant, output_dir)
        gnp_pubs = PLANT_GNP_PUBLICATIONS.get(plant['display_name'], [])
        print(f"  Created: {filepath}")
        print(f"           GNP Publications: {', '.join(gnp_pubs) if gnp_pubs else 'None'}")
        generated_files.append(filepath)

    print("\n" + "=" * 80)
    print(f"COMPLETED: Generated {len(generated_files)} Excel files")
    print(f"Output directory: {output_dir}")
    print("=" * 80 + "\n")

    return generated_files


def main():
    """Main entry point"""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, 'xlsx_templates')

    generate_all_templates(output_dir)

    print("Instructions for users:")
    print("-" * 40)
    print("1. Open your plant's .xlsx file in Excel")
    print("2. Use the 'In-House Data' sheet for book/pagination data")
    print("3. Use the 'Competitor Data' sheet for competitor data")
    print("4. Use dropdown arrows to select values (green header columns)")
    print("5. Delete the example rows before submitting")
    print("6. Check the 'Reference' sheet for all valid values")
    print("7. Return the completed file for import")
    print()


if __name__ == '__main__':
    main()
