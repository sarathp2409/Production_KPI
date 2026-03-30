"""
Import Historical Pagination Data from CSV/Excel Files

This script imports past pagination data from CSV or Excel (.xlsx) files into the database.
It handles both in-house data (PaginationReport, GNPInformation, BookDetails)
and competitor data (CompetitorData, CompetitorBookDetails).

Usage:
    python manage.py shell < pagination_dash/import_pagination_history.py

Or run in Django shell:
    exec(open('pagination_dash/import_pagination_history.py').read())

Supported File Formats:
    - CSV files: {plant_name}_inhouse.csv, {plant_name}_competitor.csv
    - Excel files: {plant_name}_pagination_data.xlsx (with sheets "In-House Data" and "Competitor Data")

Supported Fields:
    In-house data:
        - plant_id, plant_name, issue_date (required)
        - book_name, snp_pages, gnp_pages_type, number_of_gnp_jackets, book_order
        - balloon_printed, has_masthead (for B-2 books only)
        - remarks_toi, remarks_others, first_time_execution_info
        - gnp_et, gnp_et_b1, gnp_et_b2, gnp_et_had_2_books, gnp_mt, gnp_mirror, gnp_nbt, gnp_tims, gnp_remarks

    Competitor data:
        - plant_id, plant_name, issue_date, competitor_name (required)
        - main_snp_pages, main_gnp_pages_type, main_number_of_gnp_jacket, main_number_of_books
        - supp_snp_pages, supp_gnp_pages_type, supp_number_of_gnp_jacket, supp_number_of_books
        - innovation_1, innovation_2, innovation_3, comment
"""

import csv
import os
import sys
from datetime import datetime, date
from django.db import transaction
from django.contrib.auth.models import User
from django.utils import timezone

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False
    print("Warning: openpyxl not installed. Excel (.xlsx) support disabled. Install with: pip install openpyxl")

# Import models
from pagination_dash.models import (
    PaginationReport,
    GNPInformation,
    BookDetails,
    CompetitorData,
    CompetitorBookDetails
)


class PaginationDataImporter:
    """Import pagination data from CSV or Excel files"""

    # Valid choices
    BOOK_NAME_CHOICES = [choice[0] for choice in BookDetails.BOOK_NAME_CHOICES]
    GNP_TYPE_CHOICES = [choice[0] for choice in BookDetails.GNP_TYPE_CHOICES]
    INNOVATION_CHOICES = [choice[0] for choice in CompetitorData.INNOVATION_CHOICES]
    BOOK_TYPE_CHOICES = ['Main', 'Supplement']

    # Valid plant IDs
    VALID_PLANT_IDS = [
        'ODBCAIR30', 'ACTIVE42', 'ODBCAIR32', 'MAIN34', 'MAIN38',
        'MAIN35', 'MAIN40', 'ODBCAIR28', 'ACTIVE44', 'MAIN36',
        'MAIN41', 'ODBCAIR33', 'ACTIVE43', 'MAIN37'
    ]

    def __init__(self, inhouse_path, competitor_path=None, created_by_username='admin', dry_run=False, xlsx_path=None):
        """
        Initialize importer

        Args:
            inhouse_path: Path to in-house data CSV (or None if using xlsx_path)
            competitor_path: Path to competitor data CSV (or None if using xlsx_path)
            created_by_username: Username for created_by field (default: 'admin')
            dry_run: If True, validate data without committing to database (default: False)
            xlsx_path: Path to Excel file (alternative to CSV files)
        """
        self.xlsx_path = xlsx_path
        self.inhouse_csv_path = inhouse_path
        self.competitor_csv_path = competitor_path
        self.dry_run = dry_run

        # Get or create user for created_by
        self.created_by, _ = User.objects.get_or_create(
            username=created_by_username,
            defaults={'is_staff': True, 'is_superuser': True}
        )

        # Statistics
        self.stats = {
            'reports_created': 0,
            'reports_skipped': 0,
            'books_created': 0,
            'competitors_created': 0,
            'competitor_books_created': 0,
            'errors': [],
            'warnings': []
        }

    @classmethod
    def from_xlsx(cls, xlsx_path, created_by_username='admin', dry_run=False):
        """
        Create importer from an Excel file

        Args:
            xlsx_path: Path to Excel file with "In-House Data" and "Competitor Data" sheets
            created_by_username: Username for created_by field
            dry_run: If True, validate data without committing to database

        Returns:
            PaginationDataImporter instance
        """
        if not XLSX_SUPPORT:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

        return cls(
            inhouse_path=None,
            competitor_path=None,
            created_by_username=created_by_username,
            dry_run=dry_run,
            xlsx_path=xlsx_path
        )

    def _read_xlsx_sheet(self, sheet_name):
        """
        Read data from an Excel sheet and return as list of dictionaries

        Args:
            sheet_name: Name of the sheet to read

        Returns:
            List of dictionaries (one per row)
        """
        if not XLSX_SUPPORT:
            raise ImportError("openpyxl is required for Excel support")

        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in {self.xlsx_path}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        # First row is headers
        headers = [str(h).strip() if h else '' for h in rows[0]]
        data = []

        for row in rows[1:]:
            # Skip empty rows
            if not any(cell for cell in row):
                continue

            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    # Convert cell value to string, handling dates and None
                    if cell is None:
                        row_dict[headers[i]] = ''
                    elif isinstance(cell, (datetime, date)):
                        row_dict[headers[i]] = cell.strftime('%Y-%m-%d')
                    else:
                        row_dict[headers[i]] = str(cell).strip()

            data.append(row_dict)

        return data

    def _read_csv_file(self, csv_path):
        """
        Read data from a CSV file and return as list of dictionaries

        Args:
            csv_path: Path to CSV file

        Returns:
            List of dictionaries (one per row)
        """
        data = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Clean up values
                cleaned_row = {k: (v.strip() if v else '') for k, v in row.items()}
                data.append(cleaned_row)
        return data

    def parse_bool(self, value):
        """Parse boolean value from CSV"""
        if not value or value.strip() == '':
            return False
        value_lower = value.strip().lower()
        return value_lower in ['yes', 'true', '1', 'y']

    def parse_date(self, date_str):
        """Parse date from CSV (YYYY-MM-DD format)"""
        try:
            return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
        except ValueError:
            raise ValueError(f"Invalid date format: {date_str}. Expected YYYY-MM-DD")

    def parse_int(self, value, default=0):
        """Parse integer value from CSV"""
        if not value or value.strip() == '':
            return default
        try:
            return int(value.strip())
        except ValueError:
            raise ValueError(f"Invalid integer: {value}")

    def validate_choice(self, value, choices, field_name):
        """Validate that value is in allowed choices"""
        if not value or value.strip() == '':
            return None
        value = value.strip()
        if value not in choices:
            raise ValueError(
                f"Invalid {field_name}: '{value}'. "
                f"Must be one of: {', '.join(choices)}"
            )
        return value

    def import_inhouse_data(self):
        """Import in-house data from CSV or Excel"""
        print(f"\n{'='*80}")
        print(f"IMPORTING IN-HOUSE DATA")
        print(f"{'='*80}\n")

        # Read data based on file type
        if self.xlsx_path:
            if not os.path.exists(self.xlsx_path):
                print(f"❌ Error: File not found: {self.xlsx_path}")
                return
            print(f"Reading from Excel: {self.xlsx_path} (sheet: 'In-House Data')")
            data = self._read_xlsx_sheet('In-House Data')
        else:
            if not os.path.exists(self.inhouse_csv_path):
                print(f"❌ Error: File not found: {self.inhouse_csv_path}")
                return
            print(f"Reading from CSV: {self.inhouse_csv_path}")
            data = self._read_csv_file(self.inhouse_csv_path)

        # Group by report
        reports_data = {}

        for row_num, row in enumerate(data, start=2):  # Start at 2 (1 is header)
            try:
                # Create report key
                plant_id = row['plant_id'].strip()
                issue_date = self.parse_date(row['issue_date'])
                report_key = (plant_id, issue_date)

                # Validate plant_id
                if plant_id not in self.VALID_PLANT_IDS:
                    raise ValueError(f"Invalid plant_id: {plant_id}")

                # Initialize report data if first time seeing this report
                if report_key not in reports_data:
                    # Handle both old and new column names for backward compatibility
                    remarks_toi = row.get('extra_miles_toi', row.get('remarks_toi', '')).strip()
                    remarks_others = row.get('extra_miles_others', row.get('remarks_others', '')).strip()
                    et_had_2_books = self.parse_bool(row.get('et_had_2_books', row.get('gnp_et_had_2_books', '')))

                    reports_data[report_key] = {
                        'report': {
                            'plant_id': plant_id,
                            'plant_name': row['plant_name'].strip(),
                            'issue_date': issue_date,
                            'remarks_toi': remarks_toi,
                            'remarks_others': remarks_others,
                            'first_time_execution_info': row.get('first_time_execution_info', '').strip(),
                        },
                        'gnp': {
                            'et': self.parse_bool(row.get('gnp_et', '')),
                            'et_b1': self.parse_bool(row.get('gnp_et_b1', '')),
                            'et_b2': self.parse_bool(row.get('gnp_et_b2', '')),
                            'et_had_2_books': et_had_2_books,
                            'mt': self.parse_bool(row.get('gnp_mt', '')),
                            'mirror': self.parse_bool(row.get('gnp_mirror', '')),
                            'nbt': self.parse_bool(row.get('gnp_nbt', '')),
                            'tims': self.parse_bool(row.get('gnp_tims', '')),
                            'remarks': row.get('gnp_remarks', '').strip(),
                        },
                        'books': []
                    }

                # Add book data
                book_name = row['book_name'].strip()
                if book_name:  # Only add if book_name is provided
                    validated_book_name = self.validate_choice(
                        book_name,
                        self.BOOK_NAME_CHOICES,
                        'book_name'
                    )

                    # Parse B-2 specific fields (handle both old and new column names)
                    balloon_printed = self.parse_bool(
                        row.get('TOI Book-2 balloon_printed', row.get('balloon_printed', ''))
                    )
                    has_masthead = self.parse_bool(
                        row.get('TOI Book-2 has_masthead', row.get('has_masthead', ''))
                    )

                    # Warn if B-2 fields used on non-B-2 books
                    if (balloon_printed or has_masthead) and validated_book_name not in ['TOI B-2', 'TIMS B-2']:
                        warning_msg = f"Row {row_num}: balloon_printed/has_masthead ignored for {validated_book_name} (only for B-2 books)"
                        self.stats['warnings'].append(warning_msg)
                        balloon_printed = False
                        has_masthead = False

                    book_data = {
                        'book_name': validated_book_name,
                        'snp_pages': self.parse_int(row.get('snp_pages', ''), 0),
                        'gnp_pages_type': self.validate_choice(
                            row.get('gnp_pages_type', ''),
                            self.GNP_TYPE_CHOICES,
                            'gnp_pages_type'
                        ),
                        'number_of_gnp_jackets': self.parse_int(row.get('number_of_gnp_jackets', ''), 0),
                        'balloon_printed': balloon_printed,
                        'has_masthead': has_masthead,
                        'order': 0,  # Default to 0 since book_order field is removed
                    }
                    reports_data[report_key]['books'].append(book_data)

            except Exception as e:
                error_msg = f"Row {row_num}: {str(e)}"
                print(f"❌ {error_msg}")
                self.stats['errors'].append(error_msg)

        # Create reports in database
        mode_str = "[DRY RUN] " if self.dry_run else ""
        print(f"\n{mode_str}Processing {len(reports_data)} unique reports...")

        for report_key, data in reports_data.items():
            plant_id, issue_date = report_key

            try:
                with transaction.atomic():
                    # Check if report already exists
                    if PaginationReport.objects.filter(
                        plant_id=plant_id,
                        issue_date=issue_date
                    ).exists():
                        print(f"⏭️  Skipped: {plant_id} - {issue_date} (already exists)")
                        self.stats['reports_skipped'] += 1
                        continue

                    if self.dry_run:
                        # Dry run - just validate and count
                        self.stats['reports_created'] += 1
                        self.stats['books_created'] += len(data['books'])
                        print(f"✅ {mode_str}Would create: {plant_id} - {issue_date} ({len(data['books'])} books)")
                    else:
                        # Create PaginationReport
                        report = PaginationReport.objects.create(
                            **data['report'],
                            created_by=self.created_by
                        )
                        self.stats['reports_created'] += 1

                        # Create GNPInformation
                        GNPInformation.objects.create(
                            report=report,
                            **data['gnp']
                        )

                        # Create BookDetails
                        for book_data in data['books']:
                            BookDetails.objects.create(
                                report=report,
                                **book_data
                            )
                            self.stats['books_created'] += 1

                        print(f"✅ Created: {plant_id} - {issue_date} ({len(data['books'])} books)")

            except Exception as e:
                error_msg = f"Report {plant_id} - {issue_date}: {str(e)}"
                print(f"❌ {error_msg}")
                self.stats['errors'].append(error_msg)

    def import_competitor_data(self):
        """Import competitor data from CSV or Excel"""
        print(f"\n{'='*80}")
        print(f"IMPORTING COMPETITOR DATA")
        print(f"{'='*80}\n")

        # Read data based on file type
        if self.xlsx_path:
            if not os.path.exists(self.xlsx_path):
                print(f"❌ Error: File not found: {self.xlsx_path}")
                return
            print(f"Reading from Excel: {self.xlsx_path} (sheet: 'Competitor Data')")
            data = self._read_xlsx_sheet('Competitor Data')
        else:
            if not os.path.exists(self.competitor_csv_path):
                print(f"❌ Error: File not found: {self.competitor_csv_path}")
                return
            print(f"Reading from CSV: {self.competitor_csv_path}")
            data = self._read_csv_file(self.competitor_csv_path)

        for row_num, row in enumerate(data, start=2):  # Start at 2 (1 is header)
            try:
                with transaction.atomic():
                    # Get report
                    plant_id = row['plant_id'].strip()
                    issue_date = self.parse_date(row['issue_date'])

                    # Validate plant_id
                    if plant_id not in self.VALID_PLANT_IDS:
                        raise ValueError(f"Invalid plant_id: {plant_id}")

                    # Find the report
                    try:
                        report = PaginationReport.objects.get(
                            plant_id=plant_id,
                            issue_date=issue_date
                        )
                    except PaginationReport.DoesNotExist:
                        raise ValueError(
                            f"Report not found for {plant_id} - {issue_date}. "
                            "Import in-house data first."
                        )

                    competitor_name = row['competitor_name'].strip()
                    if not competitor_name:
                        print(f"⏭️  Skipped row {row_num}: No competitor name")
                        continue

                    # Check if competitor already exists
                    if CompetitorData.objects.filter(
                        report=report,
                        competitor_name=competitor_name
                    ).exists():
                        print(f"⏭️  Skipped: {plant_id} - {issue_date} - {competitor_name} (already exists)")
                        continue

                    # Parse innovation fields
                    innovations = []
                    for i in range(1, 4):  # innovation_1, innovation_2, innovation_3
                        innovation = row.get(f'innovation_{i}', '').strip()
                        if innovation:
                            validated = self.validate_choice(
                                innovation,
                                self.INNOVATION_CHOICES,
                                f'innovation_{i}'
                            )
                            if validated:
                                innovations.append(validated)

                    mode_str = "[DRY RUN] " if self.dry_run else ""

                    if self.dry_run:
                        # Dry run - just validate and count
                        self.stats['competitors_created'] += 1
                        self.stats['competitor_books_created'] += 2  # Main + Supplement
                        print(f"✅ {mode_str}Would create: {plant_id} - {issue_date} - {competitor_name}")
                    else:
                        # Create CompetitorData
                        competitor = CompetitorData.objects.create(
                            report=report,
                            competitor_name=competitor_name,
                            innovation=innovations,
                            comment=row.get('comment', '').strip()
                        )
                        self.stats['competitors_created'] += 1

                        # Create Main book
                        main_book = CompetitorBookDetails.objects.create(
                            competitor=competitor,
                            book_type='Main',
                            snp_pages=self.parse_int(row.get('main_snp_pages', ''), 0),
                            gnp_pages_type=self.validate_choice(
                                row.get('main_gnp_pages_type', ''),
                                self.GNP_TYPE_CHOICES,
                                'main_gnp_pages_type'
                            ),
                            number_of_gnp_jacket=self.parse_int(row.get('main_number_of_gnp_jacket', ''), 0),
                            number_of_books=self.parse_int(row.get('main_number_of_books', ''), 0)
                        )
                        self.stats['competitor_books_created'] += 1

                        # Create Supplement book
                        supp_book = CompetitorBookDetails.objects.create(
                            competitor=competitor,
                            book_type='Supplement',
                            snp_pages=self.parse_int(row.get('supp_snp_pages', ''), 0),
                            gnp_pages_type=self.validate_choice(
                                row.get('supp_gnp_pages_type', ''),
                                self.GNP_TYPE_CHOICES,
                                'supp_gnp_pages_type'
                            ),
                            number_of_gnp_jacket=self.parse_int(row.get('supp_number_of_gnp_jacket', ''), 0),
                            number_of_books=self.parse_int(row.get('supp_number_of_books', ''), 0)
                        )
                        self.stats['competitor_books_created'] += 1

                        print(f"✅ Created: {plant_id} - {issue_date} - {competitor_name}")

            except Exception as e:
                error_msg = f"Row {row_num}: {str(e)}"
                print(f"❌ {error_msg}")
                self.stats['errors'].append(error_msg)

    def print_summary(self):
        """Print import summary statistics"""
        mode_str = "[DRY RUN] " if self.dry_run else ""
        print(f"\n{'='*80}")
        print(f"{mode_str}IMPORT SUMMARY")
        print(f"{'='*80}\n")

        action_word = "Would create" if self.dry_run else "Created"

        print(f"📊 In-House Data:")
        print(f"   - Reports {action_word}: {self.stats['reports_created']}")
        print(f"   - Reports Skipped (duplicates): {self.stats['reports_skipped']}")
        print(f"   - Books {action_word}: {self.stats['books_created']}")

        print(f"\n📊 Competitor Data:")
        print(f"   - Competitors {action_word}: {self.stats['competitors_created']}")
        print(f"   - Competitor Books {action_word}: {self.stats['competitor_books_created']}")

        if self.stats['warnings']:
            print(f"\n⚠️  Warnings ({len(self.stats['warnings'])}):")
            for warning in self.stats['warnings'][:10]:  # Show first 10 warnings
                print(f"   - {warning}")
            if len(self.stats['warnings']) > 10:
                print(f"   ... and {len(self.stats['warnings']) - 10} more warnings")

        if self.stats['errors']:
            print(f"\n❌ Errors ({len(self.stats['errors'])}):")
            for error in self.stats['errors'][:10]:  # Show first 10 errors
                print(f"   - {error}")
            if len(self.stats['errors']) > 10:
                print(f"   ... and {len(self.stats['errors']) - 10} more errors")
        else:
            print(f"\n✅ No errors!")

        if self.dry_run:
            print(f"\n📝 This was a DRY RUN - no data was actually imported.")
            print(f"   Run again without dry_run=True to import data.")

        print(f"\n{'='*80}\n")

    def run(self):
        """Run the import process"""
        mode_str = "[DRY RUN] " if self.dry_run else ""
        print(f"\n{'='*80}")
        print(f"{mode_str}PAGINATION DATA IMPORT TOOL")
        print(f"{'='*80}\n")
        print(f"Started at: {timezone.now()}")
        print(f"Created by user: {self.created_by.username}")
        if self.dry_run:
            print(f"Mode: DRY RUN (no data will be committed)")

        # Import in-house data first
        self.import_inhouse_data()

        # Then import competitor data
        self.import_competitor_data()

        # Print summary
        self.print_summary()

        print(f"Completed at: {timezone.now()}\n")

    def run_inhouse_only(self):
        """Run only in-house data import (useful for initial import)"""
        mode_str = "[DRY RUN] " if self.dry_run else ""
        print(f"\n{'='*80}")
        print(f"{mode_str}PAGINATION DATA IMPORT TOOL (In-House Only)")
        print(f"{'='*80}\n")
        print(f"Started at: {timezone.now()}")
        print(f"Created by user: {self.created_by.username}")

        self.import_inhouse_data()
        self.print_summary()

        print(f"Completed at: {timezone.now()}\n")

    def run_competitor_only(self):
        """Run only competitor data import (requires in-house data to exist)"""
        mode_str = "[DRY RUN] " if self.dry_run else ""
        print(f"\n{'='*80}")
        print(f"{mode_str}PAGINATION DATA IMPORT TOOL (Competitor Only)")
        print(f"{'='*80}\n")
        print(f"Started at: {timezone.now()}")
        print(f"Created by user: {self.created_by.username}")

        self.import_competitor_data()
        self.print_summary()

        print(f"Completed at: {timezone.now()}\n")


# Main execution
if __name__ == '__main__':
    # Define CSV file paths - update these to point to your plant's CSV files
    base_dir = os.path.dirname(os.path.abspath(__file__))
    csv_templates_dir = os.path.join(base_dir, 'csv_templates')

    # Example: Import data for Airoli plant
    plant_name = 'Airoli'  # Change this to your plant name
    inhouse_csv = os.path.join(csv_templates_dir, f'{plant_name}_inhouse.csv')
    competitor_csv = os.path.join(csv_templates_dir, f'{plant_name}_competitor.csv')

    # Create importer and run with dry_run first to validate
    importer = PaginationDataImporter(
        inhouse_csv_path=inhouse_csv,
        competitor_csv_path=competitor_csv,
        created_by_username='admin',  # Change this if needed
        dry_run=True  # Set to False to actually import data
    )

    importer.run()

    print("✅ Import completed! You can now view the data in the admin panel or application.")


# For interactive usage in Django shell:
"""
To run this script in Django shell:

1. Start Django shell:
   python manage.py shell

2. Execute the script:
   exec(open('pagination_dash/import_pagination_history.py').read())

3. Or with custom CSV paths and dry_run mode for validation:
   from pagination_dash.import_pagination_history import PaginationDataImporter

   # First, do a dry run to validate:
   importer = PaginationDataImporter(
       inhouse_csv_path='pagination_dash/csv_templates/Airoli_inhouse.csv',
       competitor_csv_path='pagination_dash/csv_templates/Airoli_competitor.csv',
       created_by_username='your_username',
       dry_run=True  # Validate without committing
   )
   importer.run()

   # If validation passes, run again with dry_run=False to import:
   importer = PaginationDataImporter(
       inhouse_csv_path='pagination_dash/csv_templates/Airoli_inhouse.csv',
       competitor_csv_path='pagination_dash/csv_templates/Airoli_competitor.csv',
       created_by_username='your_username',
       dry_run=False  # Actually import data
   )
   importer.run()

4. To import only in-house data (without competitors):
   importer.run_inhouse_only()

5. To import only competitor data (after in-house is imported):
   importer.run_competitor_only()
"""
