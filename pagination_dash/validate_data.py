"""
Data Validation Utility for Pagination Data Import

This script validates both CSV and Excel (.xlsx) files before import to catch errors early.
It checks:
- Required fields are present
- Date formats are correct
- Choice fields contain valid values
- No duplicate entries (same date + book_name for in-house, same date + competitor for competitor)
- Plant IDs are valid

Usage:
    # Validate Excel file
    python pagination_dash/validate_data.py xlsx_templates/Airoli_pagination_data.xlsx

    # Validate CSV files
    python pagination_dash/validate_data.py csv_templates/Airoli_inhouse.csv
    python pagination_dash/validate_data.py csv_templates/Airoli_competitor.csv --type competitor

Or in Django shell:
    from pagination_dash.validate_data import DataValidator
    validator = DataValidator()
    validator.validate_excel('path/to/file.xlsx')
    validator.validate_inhouse_csv('path/to/inhouse.csv')
    validator.validate_competitor_csv('path/to/competitor.csv')
"""

import csv
import os
import sys
from datetime import datetime, date
from collections import defaultdict

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False
    print("Warning: openpyxl not installed. Excel (.xlsx) support disabled. Install with: pip install openpyxl")


class DataValidator:
    """Validate pagination CSV or Excel files before import"""

    # Valid choices
    BOOK_NAME_CHOICES = [
        'TOI B-1',
        'TOI B-2',
        'TIMS B-1',
        'TIMS B-2',
        'TOI Supp-1',
        'TOI Supp-2',
        'TOI Supp-3',
    ]

    GNP_TYPE_CHOICES = [
        '2 Pg Jkt',
        '4 Pg Jkt - 2 diff clients',
        '4 Pg Vantage',
        '4 Pg Power Jkt',
        '8 Pg Super-pano',
        '6 Pg GNP',
        '8 Pg GNP',
    ]

    INNOVATION_CHOICES = [
        'Center Panorama',
        'Bookmark',
        'Seamless Panorama',
        'Front Vantage',
        'Reverse Flap',
        'Fragrance',
        'French-Window',
        'Power Jacket',
        'Others',
    ]

    VALID_PLANT_IDS = [
        'ODBCAIR30', 'ACTIVE42', 'ODBCAIR32', 'MAIN34', 'MAIN38',
        'MAIN35', 'MAIN40', 'ODBCAIR28', 'ACTIVE44', 'MAIN36',
        'MAIN41', 'ODBCAIR33', 'ACTIVE43', 'MAIN37'
    ]

    # Required headers for in-house data
    INHOUSE_REQUIRED_HEADERS = [
        'plant_id',
        'plant_name',
        'issue_date',
        'book_name',
    ]

    # Required headers for competitor data
    COMPETITOR_REQUIRED_HEADERS = [
        'plant_id',
        'plant_name',
        'issue_date',
        'competitor_name',
    ]

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.info = []

    def reset(self):
        """Reset validation state"""
        self.errors = []
        self.warnings = []
        self.info = []

    def _read_xlsx_sheet(self, xlsx_path, sheet_name):
        """Read data from an Excel sheet and return as list of dictionaries"""
        if not XLSX_SUPPORT:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        # First row is headers
        headers = [str(h).strip() if h else '' for h in rows[0]]
        data = []

        for row in rows[1:]:
            # Skip empty rows
            if not any(cell for cell in row):
                continue

            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    # Convert cell value to string, handling dates and None
                    if cell is None:
                        row_dict[headers[i]] = ''
                    elif isinstance(cell, (datetime, date)):
                        row_dict[headers[i]] = cell.strftime('%Y-%m-%d')
                    else:
                        row_dict[headers[i]] = str(cell).strip()

            data.append(row_dict)

        return data

    def _read_csv_file(self, csv_path):
        """Read data from a CSV file and return as list of dictionaries"""
        data = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Clean up values
                cleaned_row = {k: (v.strip() if v else '') for k, v in row.items()}
                data.append(cleaned_row)
        return data

    def parse_date(self, date_str, row_num):
        """Validate and parse date"""
        if not date_str or date_str.strip() == '':
            self.errors.append(f"Row {row_num}: Missing issue_date")
            return None
        try:
            # Handle both string dates and date objects
            if isinstance(date_str, (datetime, date)):
                return date_str if isinstance(date_str, date) else date_str.date()
            return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
        except ValueError:
            self.errors.append(f"Row {row_num}: Invalid date format '{date_str}'. Expected YYYY-MM-DD")
            return None

    def validate_choice(self, value, choices, field_name, row_num, required=False):
        """Validate that value is in allowed choices"""
        if not value or value.strip() == '':
            if required:
                self.errors.append(f"Row {row_num}: Missing required field '{field_name}'")
            return None

        value = value.strip()
        if value not in choices:
            self.errors.append(
                f"Row {row_num}: Invalid {field_name} '{value}'. "
                f"Must be one of: {', '.join(choices)}"
            )
            return None
        return value

    def validate_int(self, value, field_name, row_num, allow_negative=False):
        """Validate integer value"""
        if not value or value.strip() == '':
            return 0  # Default to 0 for optional numeric fields

        try:
            int_val = int(value.strip())
            if not allow_negative and int_val < 0:
                self.warnings.append(f"Row {row_num}: Negative value for {field_name}: {int_val}")
            return int_val
        except ValueError:
            self.errors.append(f"Row {row_num}: Invalid integer for {field_name}: '{value}'")
            return None

    def validate_bool(self, value, field_name, row_num):
        """Validate boolean value"""
        if not value or value.strip() == '':
            return False

        value_lower = value.strip().lower()
        valid_true = ['yes', 'true', '1', 'y']
        valid_false = ['no', 'false', '0', 'n', '']

        if value_lower not in valid_true + valid_false:
            self.warnings.append(
                f"Row {row_num}: Unusual boolean value for {field_name}: '{value}'. "
                f"Expected: yes/no, true/false, 1/0"
            )
        return value_lower in valid_true

    def _validate_inhouse_data(self, data, source_description):
        """Validate in-house data (from CSV or Excel sheet)"""
        self.reset()
        print(f"\n{'='*80}")
        print(f"VALIDATING IN-HOUSE DATA: {source_description}")
        print(f"{'='*80}\n")

        if not data:
            self.errors.append("No data found")
            self._print_results()
            return False

        # Check required headers
        headers = list(data[0].keys()) if data else []
        for header in self.INHOUSE_REQUIRED_HEADERS:
            if header not in headers:
                self.errors.append(f"Missing required header: {header}")

        if self.errors:
            self._print_results()
            return False

        # Track duplicates
        seen_books = defaultdict(set)  # {(plant_id, issue_date): set(book_names)}
        row_count = 0
        unique_dates = set()
        unique_plants = set()

        for row_num, row in enumerate(data, start=2):
            row_count += 1

            # Validate plant_id
            plant_id = row.get('plant_id', '').strip()
            if not plant_id:
                self.errors.append(f"Row {row_num}: Missing plant_id")
            elif plant_id not in self.VALID_PLANT_IDS:
                self.errors.append(f"Row {row_num}: Invalid plant_id '{plant_id}'")
            else:
                unique_plants.add(plant_id)

            # Validate plant_name
            plant_name = row.get('plant_name', '').strip()
            if not plant_name:
                self.errors.append(f"Row {row_num}: Missing plant_name")

            # Validate issue_date
            issue_date = self.parse_date(row.get('issue_date', ''), row_num)
            if issue_date:
                unique_dates.add(issue_date)

            # Validate book_name
            book_name = self.validate_choice(
                row.get('book_name', ''),
                self.BOOK_NAME_CHOICES,
                'book_name',
                row_num,
                required=True
            )

            # Check for duplicates
            if plant_id and issue_date and book_name:
                report_key = (plant_id, issue_date)
                if book_name in seen_books[report_key]:
                    self.errors.append(
                        f"Row {row_num}: Duplicate book '{book_name}' for {plant_id} on {issue_date}"
                    )
                else:
                    seen_books[report_key].add(book_name)

            # Validate snp_pages
            self.validate_int(row.get('snp_pages', ''), 'snp_pages', row_num)

            # Validate gnp_pages_type
            gnp_type = row.get('gnp_pages_type', '').strip()
            if gnp_type:
                self.validate_choice(gnp_type, self.GNP_TYPE_CHOICES, 'gnp_pages_type', row_num)

            # Validate number_of_gnp_jackets
            self.validate_int(row.get('number_of_gnp_jackets', ''), 'number_of_gnp_jackets', row_num)

            # Validate B-2 specific fields (handle both old and new column names)
            balloon_field = 'TOI Book-2 balloon_printed' if 'TOI Book-2 balloon_printed' in row else 'balloon_printed'
            masthead_field = 'TOI Book-2 has_masthead' if 'TOI Book-2 has_masthead' in row else 'has_masthead'

            if book_name in ['TOI B-2', 'TIMS B-2']:
                self.validate_bool(row.get(balloon_field, ''), balloon_field, row_num)
                self.validate_bool(row.get(masthead_field, ''), masthead_field, row_num)
            else:
                # Warn if B-2 fields used on non-B-2 books
                if row.get(balloon_field, '').strip():
                    balloon = self.validate_bool(row.get(balloon_field, ''), balloon_field, row_num)
                    if balloon:
                        self.warnings.append(
                            f"Row {row_num}: {balloon_field} will be ignored for {book_name} (only for B-2 books)"
                        )
                if row.get(masthead_field, '').strip():
                    masthead = self.validate_bool(row.get(masthead_field, ''), masthead_field, row_num)
                    if masthead:
                        self.warnings.append(
                            f"Row {row_num}: {masthead_field} will be ignored for {book_name} (only for B-2 books)"
                        )

        # Summary info
        self.info.append(f"Total rows: {row_count}")
        self.info.append(f"Unique dates: {len(unique_dates)}")
        self.info.append(f"Unique plants: {len(unique_plants)}")
        self.info.append(f"Unique reports (plant+date): {len(seen_books)}")

        self._print_results()
        return len(self.errors) == 0

    def _validate_competitor_data(self, data, source_description):
        """Validate competitor data (from CSV or Excel sheet)"""
        self.reset()
        print(f"\n{'='*80}")
        print(f"VALIDATING COMPETITOR DATA: {source_description}")
        print(f"{'='*80}\n")

        if not data:
            self.errors.append("No data found")
            self._print_results()
            return False

        # Check required headers
        headers = list(data[0].keys()) if data else []
        for header in self.COMPETITOR_REQUIRED_HEADERS:
            if header not in headers:
                self.errors.append(f"Missing required header: {header}")

        if self.errors:
            self._print_results()
            return False

        # Track duplicates
        seen_competitors = defaultdict(set)  # {(plant_id, issue_date): set(competitor_names)}
        row_count = 0
        unique_dates = set()
        unique_plants = set()

        for row_num, row in enumerate(data, start=2):
            row_count += 1

            # Validate plant_id
            plant_id = row.get('plant_id', '').strip()
            if not plant_id:
                self.errors.append(f"Row {row_num}: Missing plant_id")
            elif plant_id not in self.VALID_PLANT_IDS:
                self.errors.append(f"Row {row_num}: Invalid plant_id '{plant_id}'")
            else:
                unique_plants.add(plant_id)

            # Validate plant_name
            plant_name = row.get('plant_name', '').strip()
            if not plant_name:
                self.errors.append(f"Row {row_num}: Missing plant_name")

            # Validate issue_date
            issue_date = self.parse_date(row.get('issue_date', ''), row_num)
            if issue_date:
                unique_dates.add(issue_date)

            # Validate competitor_name
            competitor_name = row.get('competitor_name', '').strip()
            if not competitor_name:
                self.errors.append(f"Row {row_num}: Missing competitor_name")

            # Check for duplicates
            if plant_id and issue_date and competitor_name:
                report_key = (plant_id, issue_date)
                if competitor_name in seen_competitors[report_key]:
                    self.errors.append(
                        f"Row {row_num}: Duplicate competitor '{competitor_name}' for {plant_id} on {issue_date}"
                    )
                else:
                    seen_competitors[report_key].add(competitor_name)

            # Validate Main book fields
            self.validate_int(row.get('main_snp_pages', ''), 'main_snp_pages', row_num)
            main_gnp_type = row.get('main_gnp_pages_type', '').strip()
            if main_gnp_type:
                self.validate_choice(main_gnp_type, self.GNP_TYPE_CHOICES, 'main_gnp_pages_type', row_num)
            self.validate_int(row.get('main_number_of_gnp_jacket', ''), 'main_number_of_gnp_jacket', row_num)
            self.validate_int(row.get('main_number_of_books', ''), 'main_number_of_books', row_num)

            # Validate Supplement book fields
            self.validate_int(row.get('supp_snp_pages', ''), 'supp_snp_pages', row_num)
            supp_gnp_type = row.get('supp_gnp_pages_type', '').strip()
            if supp_gnp_type:
                self.validate_choice(supp_gnp_type, self.GNP_TYPE_CHOICES, 'supp_gnp_pages_type', row_num)
            self.validate_int(row.get('supp_number_of_gnp_jacket', ''), 'supp_number_of_gnp_jacket', row_num)
            self.validate_int(row.get('supp_number_of_books', ''), 'supp_number_of_books', row_num)

            # Validate innovation fields
            for i in range(1, 4):
                innovation = row.get(f'innovation_{i}', '').strip()
                if innovation:
                    self.validate_choice(innovation, self.INNOVATION_CHOICES, f'innovation_{i}', row_num)

        # Summary info
        self.info.append(f"Total rows: {row_count}")
        self.info.append(f"Unique dates: {len(unique_dates)}")
        self.info.append(f"Unique plants: {len(unique_plants)}")
        self.info.append(f"Unique competitor entries: {sum(len(v) for v in seen_competitors.values())}")

        self._print_results()
        return len(self.errors) == 0

    def validate_excel(self, xlsx_path):
        """Validate an Excel file with both In-House and Competitor sheets"""
        if not XLSX_SUPPORT:
            print("ERROR: openpyxl is not installed. Install with: pip install openpyxl")
            return False

        if not os.path.exists(xlsx_path):
            print(f"ERROR: File not found: {xlsx_path}")
            return False

        print(f"\n{'='*80}")
        print(f"VALIDATING EXCEL FILE: {xlsx_path}")
        print(f"{'='*80}\n")

        # Validate In-House Data sheet
        try:
            inhouse_data = self._read_xlsx_sheet(xlsx_path, 'In-House Data')
            inhouse_valid = self._validate_inhouse_data(inhouse_data, f"{xlsx_path} (In-House Data sheet)")
        except ValueError as e:
            print(f"ERROR: {e}")
            inhouse_valid = False

        # Validate Competitor Data sheet
        try:
            competitor_data = self._read_xlsx_sheet(xlsx_path, 'Competitor Data')
            competitor_valid = self._validate_competitor_data(competitor_data, f"{xlsx_path} (Competitor Data sheet)")
        except ValueError as e:
            print(f"ERROR: {e}")
            competitor_valid = False

        overall_valid = inhouse_valid and competitor_valid

        if overall_valid:
            print(f"\n{'='*80}")
            print(f"✅ OVERALL VALIDATION PASSED - Excel file is ready for import")
            print(f"{'='*80}\n")
        else:
            print(f"\n{'='*80}")
            print(f"❌ OVERALL VALIDATION FAILED - Please fix errors before importing")
            print(f"{'='*80}\n")

        return overall_valid

    def validate_inhouse_csv(self, csv_path):
        """Validate in-house data CSV"""
        if not os.path.exists(csv_path):
            print(f"ERROR: File not found: {csv_path}")
            return False

        data = self._read_csv_file(csv_path)
        return self._validate_inhouse_data(data, csv_path)

    def validate_competitor_csv(self, csv_path):
        """Validate competitor data CSV"""
        if not os.path.exists(csv_path):
            print(f"ERROR: File not found: {csv_path}")
            return False

        data = self._read_csv_file(csv_path)
        return self._validate_competitor_data(data, csv_path)

    def _print_results(self):
        """Print validation results"""
        print("\n📊 Summary:")
        for info in self.info:
            print(f"   {info}")

        if self.warnings:
            print(f"\n⚠️  Warnings ({len(self.warnings)}):")
            for warning in self.warnings[:20]:
                print(f"   - {warning}")
            if len(self.warnings) > 20:
                print(f"   ... and {len(self.warnings) - 20} more warnings")

        if self.errors:
            print(f"\n❌ Errors ({len(self.errors)}):")
            for error in self.errors[:30]:
                print(f"   - {error}")
            if len(self.errors) > 30:
                print(f"   ... and {len(self.errors) - 30} more errors")
            print(f"\n❌ VALIDATION FAILED - Please fix errors before importing")
        else:
            print(f"\n✅ VALIDATION PASSED - Data is ready for import")

        print(f"\n{'='*80}\n")


def main():
    """Main entry point for command line usage"""
    import argparse

    parser = argparse.ArgumentParser(description='Validate pagination CSV or Excel files')
    parser.add_argument('file_path', help='Path to file to validate (.csv or .xlsx)')
    parser.add_argument(
        '--type', '-t',
        choices=['inhouse', 'competitor', 'excel', 'auto'],
        default='auto',
        help='Type of file (default: auto-detect from filename/extension)'
    )

    args = parser.parse_args()

    validator = DataValidator()

    # Auto-detect type from filename if not specified
    file_type = args.type
    filename = os.path.basename(args.file_path).lower()
    file_ext = os.path.splitext(filename)[1]

    if file_type == 'auto':
        if file_ext == '.xlsx':
            file_type = 'excel'
        elif 'competitor' in filename:
            file_type = 'competitor'
        elif 'inhouse' in filename:
            file_type = 'inhouse'
        else:
            print("Could not auto-detect file type from filename.")
            print("Please specify --type inhouse, --type competitor, or --type excel")
            sys.exit(1)

    if file_type == 'excel':
        success = validator.validate_excel(args.file_path)
    elif file_type == 'inhouse':
        success = validator.validate_inhouse_csv(args.file_path)
    else:
        success = validator.validate_competitor_csv(args.file_path)

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
